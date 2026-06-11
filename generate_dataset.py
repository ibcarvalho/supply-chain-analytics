import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

pedidos      = pd.read_csv('data/pedidos_compra.csv')
fornecedores = pd.read_csv('data/fornecedores.csv')

AZ1='1F3864'; AZ2='2E75B6'; GR1='1E8449'; GR2='D5F5E3'
RD1='C0392B'; RD2='FADBD8'; AM1='9A7D0A'; AM2='FEF9E7'
CZ1='404040'; CZ2='F2F2F2'; BR='FFFFFF'

def borda():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

def hdr(ws,r,c,v,bg=AZ1,fg=BR,sz=10,wrap=False,al='center'):
    x=ws.cell(r,c,v)
    x.font=Font(bold=True,color=fg,size=sz,name='Arial')
    x.fill=PatternFill('solid',fgColor=bg)
    x.alignment=Alignment(horizontal=al,vertical='center',wrap_text=wrap)
    x.border=borda(); return x

def cel(ws,r,c,v=None,fmt=None,bg=BR,bold=False,al='center',fg=CZ1):
    x=ws.cell(r,c,v)
    x.font=Font(bold=bold,color=fg,size=10,name='Arial')
    x.fill=PatternFill('solid',fgColor=bg)
    x.alignment=Alignment(horizontal=al,vertical='center')
    x.border=borda()
    if fmt: x.number_format=fmt
    return x

def fcl(ws,r,c,f,fmt=None,bg=BR,bold=False,al='center',fg='000000'):
    x=ws.cell(r,c,f)
    x.font=Font(bold=bold,color=fg,size=10,name='Arial')
    x.fill=PatternFill('solid',fgColor=bg)
    x.alignment=Alignment(horizontal=al,vertical='center')
    x.border=borda()
    if fmt: x.number_format=fmt
    return x

def titulo(ws,r,txt,cols=12):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=cols)
    x=ws.cell(r,1,txt)
    x.font=Font(bold=True,color=BR,size=13,name='Arial')
    x.fill=PatternFill('solid',fgColor=AZ1)
    x.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[r].height=32

def subtitulo(ws,r,txt,cols=12,bg=AZ2):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=cols)
    x=ws.cell(r,1,txt)
    x.font=Font(bold=True,color=BR,size=11,name='Arial')
    x.fill=PatternFill('solid',fgColor=bg)
    x.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    ws.row_dimensions[r].height=22

def nota(ws,r,txt,cols=10,bg='EEF2FF'):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=cols)
    x=ws.cell(r,1,txt)
    x.font=Font(italic=True,size=9,color='444455',name='Arial')
    x.fill=PatternFill('solid',fgColor=bg)
    x.alignment=Alignment(wrap_text=True,indent=1)
    ws.row_dimensions[r].height=28

def larguras(ws,d):
    for c,w in d.items(): ws.column_dimensions[c].width=w

wb=Workbook()

# 
# ABA FORNECEDORES 
# ════════════════════════════════════════════════════════
ws_f=wb.active; ws_f.title='Fornecedores'
ws_f.sheet_view.showGridLines=False


cols_f=[
    ('id_fornecedor','ID'),
    ('nome_fornecedor','Fornecedor'),
    ('cnpj','CNPJ'),
    ('categoria','Categoria'),
    ('cidade','Cidade'),
    ('contrato_ativo','Ativo'),
    ('prazo_entrega_dias','Prazo Dias'),   
    ('avaliacao','Avaliacao'),             
]
for i,(_,lb) in enumerate(cols_f,1): hdr(ws_f,1,i,lb)
ws_f.row_dimensions[1].height=22

for ri,row in enumerate(fornecedores[[c for c,_ in cols_f]].itertuples(index=False),2):
    bg=BR if ri%2==0 else CZ2
    for ci,v in enumerate(row,1):
        c=ws_f.cell(ri,ci,v)
        c.font=Font(size=10,name='Arial',color=CZ1)
        c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='left' if ci==2 else 'center',vertical='center')
        c.border=borda()
        if ci==8: c.number_format='0.0'

nf=len(fornecedores)
tf=Table(displayName='tForn',ref=f'A1:{get_column_letter(len(cols_f))}{nf+1}')
tf.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True)
ws_f.add_table(tf)
larguras(ws_f,{'A':6,'B':22,'C':22,'D':22,'E':16,'F':8,'G':12,'H':12})

# ABA PEDIDOS — nomes de coluna SEM acentos/simbolos
# # =======================================================================
ws_p=wb.create_sheet('Pedidos')
ws_p.sheet_view.showGridLines=False

cols_p=[
    ('id_pedido',           'ID Pedido'),
    ('id_produto',          'ID Produto'),
    ('nome_produto',        'Produto'),
    ('id_fornecedor',       'ID Forn'),     
    ('nome_fornecedor',     'Fornecedor'),
    ('categoria',           'Categoria'),
    ('data_solicitacao',    'Dt Solic'),
    ('data_entrega_prevista','Dt Prev'),
    ('data_entrega_real',   'Dt Real'),
    ('quantidade',          'Qtde'),
    ('preco_unitario',      'Preco Unit'),  
    ('valor_total',         'Valor Total'),
    ('status',              'Status'),
    ('dias_atraso',         'Dias Atraso'),
    ('lead_time_real',      'Lead Time'),
    ('mes',                 'Mes Num'),     
    ('mes_nome',            'Mes Nome'),
    ('trimestre',           'Trimestre'),
    ('semestre',            'Semestre'),
]
for i,(_,lb) in enumerate(cols_p,1): hdr(ws_p,1,i,lb)
ws_p.row_dimensions[1].height=22

for ri,row in enumerate(pedidos[[c for c,_ in cols_p]].itertuples(index=False),2):
    bg=BR if ri%2==0 else CZ2
    for ci,v in enumerate(row,1):
        vv='' if (v is None or (isinstance(v,float) and np.isnan(v))) else v
        c=ws_p.cell(ri,ci,vv)
        c.font=Font(size=9,name='Arial',color=CZ1)
        c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=borda()
        if ci==11: c.number_format='R$ #,##0.00'
        if ci==12: c.number_format='R$ #,##0.00'

np_=len(pedidos)
tp=Table(displayName='tPed',ref=f'A1:{get_column_letter(len(cols_p))}{np_+1}')
tp.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True)
ws_p.add_table(tp)
larguras(ws_p,{'A':15,'B':10,'C':26,'D':9,'E':20,'F':18,
               'G':13,'H':13,'I':13,'J':7,'K':14,'L':16,
               'M':10,'N':12,'O':12,'P':9,'Q':10,'R':10,'S':10})
# =======================================================================
wb.save('analise_suprimentos.xlsx')
print("Salvo!")